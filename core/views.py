#from django.shortcuts import render
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, FormView
from core import models, forms
from django.contrib.staticfiles.templatetags.staticfiles import static
from django.http import HttpResponseRedirect, HttpResponse

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Protection, Font, Style

from openpyxl.cell import get_column_letter

from django.http import JsonResponse
from django.views.generic.edit import CreateView
from io import BytesIO

from django.core.urlresolvers import reverse

class AjaxableResponseMixin(object):
    """
    Mixin to add AJAX support to a form.
    Must be used with an object-based FormView (e.g. CreateView)
    """
    def form_invalid(self, form):
        response = super(AjaxableResponseMixin, self).form_invalid(form)
        if self.request.is_ajax():
            return JsonResponse(form.errors, status=400)
        else:
            return response

    def form_valid(self, form):
        # A custom function which should be present in all data adding forms
        xlsx_with_errors = form.process_data(project_pk=self.kwargs['project_pk'])

        # We make sure to call the parent's form_valid() method because
        # it might do some processing (in the case of CreateView, it will
        # call form.save() for example).
        # response = super(AjaxableResponseMixin, self).form_valid(form)
        # removed the above as I don't think i need it for my custom function
        import pdb; pdb.set_trace()
        if self.request.is_ajax():
            print('returning data')
            data = {
                #'pk': self.object.pk,
                'error_sheet': xlsx_with_errors
            }
            return JsonResponse(data)
        else:
            return HttpResponseRedirect(self.get_success_url())


def population_data_spreadsheet_validation(wb, ws, species_validation_sheet, genus_validation_sheet):
    # Lock the validation sheets so they cannot be tampered with
    species_validation_sheet.protection.enable()
    genus_validation_sheet.protection.enable()

    # Add the protect to the sheet, and remove it for the individual cells
    ws.protection.sheet = True
    for row in ws.iter_rows('A2:F2000'):
        for cell in row:
            cell.style = Style(protection=Protection(locked=False, hidden=False))

    # Create & add the validation
    genera_dv = DataValidation(
        type='list',
        formula1="='Valid Genera'!A1:A" + str(genus_validation_sheet.max_row),
        error='Invalid genus. Please see the "Valid Genera" sheet to view allowed genera.',
        promptTitle='Restricted list',
        prompt='Please see the "Valid Genera" sheet to view allowed genera.'
    )
    species_dv = DataValidation(
        type='list',
        formula1="='Valid Species'!A1:A" + str(species_validation_sheet.max_row),
        error='Invalid genus. Please see the "Valid Species" sheet to view allowed species.',
        promptTitle='Restricted list',
        prompt='Please see the "Valid Species" sheet to view allowed species.'
    )
    ws.add_data_validation(genera_dv)
    ws.add_data_validation(species_dv)
    genera_dv.ranges.append('A2:A1048576')
    species_dv.ranges.append('B2:B1048576')

    # Add the list validation
    collision_risk_dv = DataValidation(
        type="list",
        formula1='"High,Medium,Low"',
        promptTitle='Restricted list',
        prompt='Please enter either: "High", "Medium" or "Low".',
        error='Invalid collision risk. Please enter either: "High", "Medium" or "Low".'
    )
    ws.add_data_validation(collision_risk_dv)
    collision_risk_dv.ranges.append('D2:D1048576')

    # Add the number validation
    count_dv = DataValidation(
        type='whole',
        operator='greaterThan',
        formula1=0,
        prompt='Please enter a whole number greater than 0.',
        error='Invalid. Please enter a whole number greater than 0.'
    )
    ws.add_data_validation(count_dv)
    count_dv.ranges.append('C2:C1048576')
    density_km_dv = DataValidation(  # operator="between", formula1=0, formula2=1
        type='decimal',
        operator='greaterThan',
        formula1=0.01,
        prompt='Please enter a number greater than 0.',
        error='Invalid. Please enter a number greater than 0.'
    )
    ws.add_data_validation(density_km_dv)
    density_km_dv.ranges.append('E1:E1048576')
    passage_rate_dv = DataValidation(
        type='whole',
        operator='greaterThan',
        formula1=0,
        prompt='Please enter a whole number greater than 0.',
        error='Invalid. Please enter a whole number greater than 0.'
    )
    ws.add_data_validation(passage_rate_dv)
    passage_rate_dv.ranges.append('F2:F1048576')


def create_population_data_spreadsheet():
    # Create the workbook
    wb = Workbook()

    # Create the template spreadsheet
    ws = wb.active
    ws.title = 'Main'
    ws.sheet_properties.tabColor = "1072BA"

    # Freeze panes
    ws.freeze_panes = ws['A2']

    # Add the columns
    ws['A1'] = 'genus'
    ws['B1'] = 'species'
    ws['C1'] = 'count'
    ws['D1'] = 'collision_risk'
    ws['E1'] = 'density_km'
    ws['F1'] = 'passage_rate'

    # Format them in bold
    heading = Style(font=Font(bold=True), protection=Protection(locked=True, hidden=False))
    for row in ws.iter_rows('A1:F1'):
        for cell in row:
            cell.style = heading

    # Get a list of the valid species & genera for validation
    genera = sorted(list(models.Taxa.objects.values_list('genus', flat=True).distinct()))
    species = sorted(list(models.Taxa.objects.values_list('species', flat=True).distinct()))

    # Create additional sheets to hold them
    genus_validation_sheet = wb.create_sheet()
    genus_validation_sheet.title = 'Valid Genera'
    species_validation_sheet = wb.create_sheet()
    species_validation_sheet.title = 'Valid Species'

    # Population the cells
    for i, g in enumerate(genera, 1):
        genus_validation_sheet.cell(row=i, column=1, value=g)
    for i, s in enumerate(species, 1):
        species_validation_sheet.cell(row=i, column=1, value=s)

    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 13
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 100

    # Add the validation
    population_data_spreadsheet_validation(wb, ws, species_validation_sheet, genus_validation_sheet)

    # Close & save
    wb.save('core' + static('population_data_for_upload.xlsx'))


class PopulationDataCreateView(FormView):#class PopulationDataCreateView(AjaxableResponseMixin, FormView):
    template_name = 'core/populationdata_create_form.html'
    form_class = forms.MetaDataCreateForm
    #create_population_data_spreadsheet()

    def get_success_url(self):
        return reverse('project_detail', args={'pk': self.kwargs['project_pk']})

    # No idea why this needs to go in, but it seems to http://stackoverflow.com/questions/18605008/curious-about-get-form-kwargs-in-formview
    def get_form_kwargs(self):
        kwargs = super(PopulationDataCreateView, self).get_form_kwargs()
        kwargs['project_pk'] = self.kwargs['project_pk']
        return kwargs

    def get_context_data(self, **kwargs):
        context = super(PopulationDataCreateView, self).get_context_data(**kwargs)
        context['project_pk'] = self.kwargs['project_pk']
        return context

    def form_invalid(self, form):
        response = super(PopulationDataCreateView, self).form_invalid(form)
        if self.request.is_ajax():
            return JsonResponse(form.errors, status=400)
        else:
            return response

    def form_valid(self, form):
        print('calling form_valid')
        # A custom function which should be present in all data adding forms
        xlsx_with_errors = form.process_data(project_pk=self.kwargs['project_pk'])

        # We make sure to call the parent's form_valid() method because
        # it might do some processing (in the case of CreateView, it will
        # call form.save() for example).
        # response = super(AjaxableResponseMixin, self).form_valid(form)
        # removed the above as I don't think i need it for my custom function

        if self.request.is_ajax():
            print('returning data')
            data = {
                #'pk': self.object.pk,
                'error_sheet': xlsx_with_errors
            }
            return JsonResponse(data)
        else:
            return HttpResponseRedirect(self.get_success_url())
    # create_population_data_spreadsheet() TODO this needs to be called through some chron thing
'''
    # This method is called when valid form data has been POSTed. It should return an HttpResponse.
    def form_valid(self, form):
        # This is taking a long time...
        print('form_valid is running')

        # Run the data processing function
        form.process_data(project_pk=self.kwargs['project_pk'])

        # I don't think we need to run the super function, as the process_data should have taken care of all that
        # return super(PopulationDataCreateView, self).form_valid(form)

        return HttpResponseRedirect(self.get_success_url())'''

class FocalSiteDataCreate(CreateView):
    model = models.FocalSiteData
    template_name_suffix = '_create_form'
    form_class = forms.FocalSiteDataCreateForm


class FocalSiteCreate(CreateView):
    model = models.FocalSite
    template_name_suffix = '_create_form'
    form_class = forms.FocalSiteCreateForm


class ProjectList(ListView):
    model = models.Project
    context_object_name = 'projects'


class ProjectDetail(DetailView):
    model = models.Project
    context_object_name = 'project'


class ProjectCreate(CreateView):
    model = models.Project
    template_name_suffix = '_create_form'
    form_class = forms.ProjectCreateForm


class ProjectUpdate(UpdateView):
    model = models.Project
    template_name_suffix = '_update_form'
    form_class = forms.ProjectUpdateForm


class ProjectDelete(DeleteView):
    model = models.Project
    template_name_suffix = '_delete_form'
    form_class = forms.ProjectDeleteForm


class DeveloperCreate(CreateView):
    model = models.Developer
    fields = ['name', 'email', 'phone']
    template_name_suffix = '_create_form'


class DataList(ListView):
    model = models.Project
    context_object_name = 'projects'


class PopulationDataCreate(CreateView):
    model = models.PopulationData
    template_name_suffix = '_create_form'
    form_class = forms.PopulationDataCreateForm


'''def create_population_data(request, project_pk):
    metadata_form = forms.MetaDataCreateForm()

    if request.POST:
            metadata_form = forms.MetaDataCreateForm(request.POST, prefix='meta')
            data_form = forms.PopulationDataCreateForm(request.POST, request.FILES, prefix='meta')

            if metadata_form.is_valid() and data_form.is_valid():
                    # Prepare & save the metadata
                    metadata = metadata_form.save(commit=False)
                    metadata.project = models.Project.objects.get(pk=project_pk)
                    metadata.save()

                    # Parse the spreadsheet and get the data
                    data = request.FILES['spreadsheet']

                    # Prepare & save the data
                    data = data_form.save(commit=False)
                    data.metadata = metadata
                    data.save()

    # Create a spreadsheet with validation for them on-the-fly - note we may need to change this if it impacts performance
    workbook = xlsxwriter.Workbook('population_data_for_upload.xlsx')
    worksheet = workbook.add_worksheet()
    worksheet.write('A1', 'Hello world')
    workbook.close()

    return render_to_response('core/populationdata_create_form.html', {
        'metadata_form': metadata_form,
        'data_form': data_form,
        'project_pk': project_pk,
        },
        RequestContext(request))'''

